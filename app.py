import os
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import razorpay
from dotenv import load_dotenv
from PyPDF2 import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from openpyxl import load_workbook
import io

load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///site.db'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB limit

db = SQLAlchemy(app)

# Initialize Razorpay client
razorpay_client = razorpay.Client(auth=(os.getenv('RAZORPAY_KEY_ID'), os.getenv('RAZORPAY_KEY_SECRET')))

# Login Manager
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# Models
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(20), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(60), nullable=False)
    subscriptions = db.relationship('Subscription', backref='user', lazy=True)
    tasks_this_month = db.Column(db.Integer, default=0)
    last_task_date = db.Column(db.DateTime)

class Subscription(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    plan_id = db.Column(db.String(50), nullable=False)
    start_date = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    end_date = db.Column(db.DateTime, nullable=False)
    devices = db.Column(db.Integer, nullable=False)
    payment_id = db.Column(db.String(100), nullable=False)
    status = db.Column(db.String(20), default='active')

# Create database tables
with app.app_context():
    db.create_all()

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Helper Functions
def get_subscription_plans():
    return {
        '1m1d': {'name': '1 Month (1 Device)', 'price': 19900, 'duration': 30, 'devices': 1},
        '1m2d': {'name': '1 Month (2 Devices)', 'price': 24900, 'duration': 30, 'devices': 2},
        '3m1d': {'name': '3 Months (1 Device)', 'price': 49900, 'duration': 90, 'devices': 1},
        '3m2d': {'name': '3 Months (2 Devices)', 'price': 54900, 'duration': 90, 'devices': 2},
        '1y5d': {'name': '1 Year (5 Devices)', 'price': 200000, 'duration': 365, 'devices': 5}
    }

def check_user_subscription(user):
    if not user:
        return False
    active_sub = Subscription.query.filter(
        Subscription.user_id == user.id,
        Subscription.end_date >= datetime.utcnow(),
        Subscription.status == 'active'
    ).first()
    return active_sub

def reset_monthly_tasks_if_new_month(user):
    if user.last_task_date and user.last_task_date.month != datetime.utcnow().month:
        user.tasks_this_month = 0
        db.session.commit()

# Routes
@app.route('/')
def index():
    user = current_user if current_user.is_authenticated else None
    is_subscribed = check_user_subscription(user)
    sub_expiry_date = is_subscribed.end_date.strftime('%d %b %Y') if is_subscribed else None
    
    return render_template('index.html', 
                         user=user.username if user else None,
                         is_subscribed=bool(is_subscribed),
                         sub_expiry_date=sub_expiry_date)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password, password):
            login_user(user)
            return redirect(url_for('index'))
        else:
            flash('Invalid username or password', 'error')
    return render_template('login.html')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        
        if User.query.filter_by(username=username).first():
            flash('Username already taken', 'error')
        elif User.query.filter_by(email=email).first():
            flash('Email already registered', 'error')
        else:
            hashed_password = generate_password_hash(password)
            new_user = User(username=username, email=email, password=hashed_password)
            db.session.add(new_user)
            db.session.commit()
            login_user(new_user)
            return redirect(url_for('index'))
    return render_template('signup.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

@app.route('/plans')
@login_required
def plans():
    plans = get_subscription_plans()
    return render_template('plans.html', plans=plans)

@app.route('/create_order', methods=['POST'])
@login_required
def create_order():
    plan_id = request.form.get('plan_id')
    plans = get_subscription_plans()
    plan = plans.get(plan_id)
    
    if not plan:
        return jsonify({'error': 'Invalid plan'}), 400
    
    data = {
        'amount': plan['price'],
        'currency': 'INR',
        'receipt': f'sub_{current_user.id}_{datetime.now().timestamp()}',
        'notes': {
            'user_id': current_user.id,
            'plan_id': plan_id
        }
    }
    order = razorpay_client.order.create(data=data)
    return jsonify(order)

@app.route('/payment_success', methods=['POST'])
@login_required
def payment_success():
    payment_id = request.form.get('razorpay_payment_id')
    order_id = request.form.get('razorpay_order_id')
    signature = request.form.get('razorpay_signature')
    plan_id = request.form.get('plan_id')
    
    try:
        # Verify payment
        params_dict = {
            'razorpay_order_id': order_id,
            'razorpay_payment_id': payment_id,
            'razorpay_signature': signature
        }
        razorpay_client.utility.verify_payment_signature(params_dict)
        
        # Get plan details
        plans = get_subscription_plans()
        plan = plans.get(plan_id)
        
        # Create subscription
        start_date = datetime.utcnow()
        end_date = start_date + timedelta(days=plan['duration'])
        
        new_sub = Subscription(
            user_id=current_user.id,
            plan_id=plan_id,
            start_date=start_date,
            end_date=end_date,
            devices=plan['devices'],
            payment_id=payment_id
        )
        db.session.add(new_sub)
        db.session.commit()
        
        return redirect(url_for('dashboard'))
    except Exception as e:
        flash('Payment verification failed', 'error')
        return redirect(url_for('plans'))

@app.route('/dashboard')
@login_required
def dashboard():
    active_sub = check_user_subscription(current_user)
    plans = get_subscription_plans()
    sub_plan = plans.get(active_sub.plan_id) if active_sub else None
    
    return render_template('dashboard.html', 
                         active_sub=active_sub,
                         sub_plan=sub_plan,
                         tasks_remaining=max(0, 2 - current_user.tasks_this_month))

@app.route('/highlight', methods=['POST'])
@login_required
def highlight_route():
    reset_monthly_tasks_if_new_month(current_user)
    active_sub = check_user_subscription(current_user)
    
    # Check task limit for free users
    if not active_sub and current_user.tasks_this_month >= 2:
        flash('Free limit reached (2 tasks/month). Please subscribe for unlimited access.', 'error')
        return redirect(url_for('plans'))
    
    # Process files
    pdf_file = request.files['pdf_file']
    excel_file = request.files['excel_file']
    highlight_type = request.form['highlight_type']
    
    # Save files temporarily
    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], 'temp.pdf')
    excel_path = os.path.join(app.config['UPLOAD_FOLDER'], 'temp.xlsx')
    pdf_file.save(pdf_path)
    excel_file.save(excel_path)
    
    # Process files
    if highlight_type == 'uan':
        result_pdf, missing_data = highlight_uan_numbers(pdf_path, excel_path)
    else:
        result_pdf, missing_data = highlight_esic_numbers(pdf_path, excel_path)
    
    # Update task count
    current_user.tasks_this_month += 1
    current_user.last_task_date = datetime.utcnow()
    db.session.commit()
    
    return render_template('result.html', 
                         pdf_data=result_pdf,
                         missing_data=missing_data,
                         highlight_type=highlight_type)

def highlight_uan_numbers(pdf_path, excel_path):
    # Extract UAN numbers from Excel (Column A)
    wb = load_workbook(excel_path)
    ws = wb.active
    uan_numbers = [str(cell.value).strip() for cell in ws['A'] if cell.value]
    
    # Process PDF
    pdf_reader = PdfReader(pdf_path)
    pdf_writer = PdfWriter()
    
    for page in pdf_reader.pages:
        packet = io.BytesIO()
        can = canvas.Canvas(packet, pagesize=letter)
        
        text = page.extract_text()
        for uan in uan_numbers:
            if uan in text:
                # Find positions of the UAN in text
                # This is simplified - you'd need proper text position detection
                can.setFillColor(colors.yellow)
                can.rect(100, 100, 100, 20, fill=1, stroke=0)  # Example position
        
        can.save()
        packet.seek(0)
        overlay = PdfReader(packet)
        page.merge_page(overlay.pages[0])
        pdf_writer.add_page(page)
    
    # Save output PDF
    output_pdf = io.BytesIO()
    pdf_writer.write(output_pdf)
    output_pdf.seek(0)
    
    # Find missing UANs
    found_uan = set()
    for page in pdf_reader.pages:
        text = page.extract_text()
        for uan in uan_numbers:
            if uan in text:
                found_uan.add(uan)
    
    missing_uan = [uan for uan in uan_numbers if uan not in found_uan]
    
    # Create missing data Excel
    missing_wb = load_workbook(excel_path)
    missing_ws = missing_wb.active
    missing_ws.append(["Missing UAN Numbers"])
    for uan in missing_uan:
        missing_ws.append([uan])
    
    missing_data = io.BytesIO()
    missing_wb.save(missing_data)
    missing_data.seek(0)
    
    return output_pdf.getvalue(), missing_data.getvalue()

def highlight_esic_numbers(pdf_path, excel_path):
    # Similar to UAN but highlight entire row
    # Implementation would be similar with different highlighting logic
    pass

# Other static pages
@app.route('/about')
def about():
    return render_template('base.html')  # You should create specific templates for these

@app.route('/refund')
def refund():
    return render_template('base.html')

@app.route('/shipping')
def shipping():
    return render_template('base.html')

@app.route('/terms')
def terms():
    return render_template('base.html')

@app.route('/privacy')
def privacy():
    return render_template('base.html')

@app.route('/contact')
def contact():
    return render_template('base.html')

if __name__ == '__main__':
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    app.run(debug=True)
